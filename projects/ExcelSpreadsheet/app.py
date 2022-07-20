# There are several Excel files in this folder to work with

import openpyxl

work_book_a = openpyxl.Workbook()
print ('work_book_a = ',work_book_a)
print ('type(work_book_a) = ',type(work_book_a))
# This instruction creates a <class 'openpyxl.workbook.workbook.Workbook'> object

work_book_b = openpyxl.load_workbook('users.xlsx')
print ('work_book_b = ',work_book_b)
print ('type(work_book_b) = ',type(work_book_b))
# This instruction creates a <class 'openpyxl.workbook.workbook.Workbook'> object loading data from a file
# passed as a constructor

sheetname_list = work_book_b.sheetnames
print ('sheetname_list = ',sheetname_list)
print ('type(sheetname_list) = ',type(sheetname_list))
# This attribute returns a list with the names of every book or spread sheet in the file

# To select a particular sheet from the workbook there has to be passed the sheet name
sheet_a = work_book_b ['Sheet_1']
print('sheet_1 = ',sheet_a)
print('type(sheet_1) = ',type(sheet_a))
# This instruction creates a <class 'openpyxl.worksheet.worksheet.Worksheet'>

cell_a = sheet_a['A1']
print ('cell_a = ',cell_a)
print ('type(cell_a) = ',type(cell_a))
# This instruction creates a <class 'openpyxl.cell.cell.Cell'>

# From a <class 'openpyxl.cell.cell.Cell'> there can be selected the rows, the columns and the values
cell_a_value = cell_a.value
print ('cell_a_value = ',cell_a_value)
print ('type(cell_a_value) = ',type(cell_a_value))
# This attribute returns the actual value of such cell

cell_a_row = cell_a.row
print ('cell_a_row = ',cell_a_row)
print ('type(cell_a_row) = ',type(cell_a_row))
# This attribute returns the row in the form of a integer starting from 1

cell_a_column = cell_a.column
print ('cell_a_column = ',cell_a_column)
print ('type(cell_a_column) = ',type(cell_a_column))
# This attribute returns the column in the form of a integer starting from 1

cell_a_coordinate = cell_a.coordinate
print ('cell_a_coordinate = ',cell_a_coordinate)
print ('type(cell_a_coordinate) = ',type(cell_a_coordinate))
# This attribute returns the coordinate in the form of a string starting from 'A1'

# There's another approeach to get a <class 'openpyxl.cell.cell.Cell'> object
cell_b = sheet_a.cell(row=1, column=1)
print ('cell_b = ',cell_b)
print ('type(cell_b) = ',type(cell_b))
# This instruction creates a <class 'openpyxl.cell.cell.Cell'>

# To get the actual number of columns and rows there's an attribute to use
rows_a = sheet_a.max_row
print ('rows_a = ',rows_a)
# This returns the max number of rows

columns_a = sheet_a.max_column
print ('columns_a = ',columns_a)
# This returns the max number of columns

# Sheets can be iterated using cell coordinates, either through columns, rows or both
for row in range (1,rows_a):
    for column in range (1,columns_a):
        cell = sheet_a.cell (row,column)
        print ('cell = ', cell,'\ncell.value = ',cell.value,'\n')

# This loop iterates each column for each row

for column in range (1,columns_a):
    for row in range (1,rows_a):
        cell = sheet_a.cell (row,column)
        print ('cell = ', cell,'\ncell.value = ',cell.value,'\n')

# This loop iterates each row for each column

# There can be accessed a whole ranage of cells in a column or in a row
column_b = sheet_a ['A']
print ('column_b = ',column_b)
print ('type(column_b) = ',type(column_b))
print ('column_b[0] = ',column_b[0])
print ('type(column_b[0]) = ',type(column_b[0]))
# This instruction creates a <class 'tuple'> containing in it's interior class 'openpyxl.cell.cell.Cell'> objects 
# for every cell in the column

rows_b = sheet_a ['1']
print ('rows_b = ',rows_b)
print ('type(rows_b) = ',type(rows_b))
print ('rows_b[0] = ',rows_b[0])
print ('type(rows_b[0]) = ',type(rows_b[0]))
# This instruction creates a <class 'tuple'> containing in it's interior class 'openpyxl.cell.cell.Cell'> objects
# for every cell in the row. Note that this mwthod allows to recieve an integer when we're specifiying rows

range_a = sheet_a ['A1:B5']
print ('range_a = ',range_a)
print ('type(range_a) = ',type(range_a))
print ('range_a[0] = ',range_a[0])
print ('type(range_a[0]) = ',type(range_a[0]))
# This instruction creates a <class 'tuple'> containing in it's interior class 'openpyxl.cell.cell.Cell'> objects
# for every cell in the range. The tuple organizes by row then column

range_b = sheet_a ['A:C']
print ('range_b = ',range_b)
print ('type(range_b) = ',type(range_b))
print ('range_b[0] = ',range_b[0])
print ('type(range_b[0]) = ',type(range_b[0]))
# This instruction creates a <class 'tuple'> containing in it's interior class 'openpyxl.cell.cell.Cell'> objects
# for every cell in the range. The tuple organizes by row then column

range_c = sheet_a ['1:3']
print ('range_c = ',range_c)
print ('type(range_c) = ',type(range_c))
print ('range_c[0] = ',range_c[0])
print ('type(range_c[0]) = ',type(range_c[0]))
# This instruction creates a <class 'tuple'> containing in it's interior class 'openpyxl.cell.cell.Cell'> objects
# for every cell in the range. The tuple organizes by row then column

# To add a row at the end of the sheet there can be employed the append method

sheet_a.append (['Eyes',21,'online'])
# This method addds values row-wise at the end of the sheet

# There are several method used in spreadsheets

sheet_a.insert_rows(3,5)
# 'insert_rows()' places at the 3rd row 5 empty rows

sheet_a.insert_cols(1,3)
# 'insert_cols()' places at the 1st row 3 empty columns

sheet_a.delete_rows(3,5)
# 'delete_rows()' deletes at the 3rd row 5 rows

sheet_a.delete_cols(1,3)
# 'delete_cols()' deletes at the 1st row 3 empty columns

# To create a new sheet on a workbook there can be used the 'create_sheet()' method
work_book_b.create_sheet('Py')
print ('work_book_b.sheetnames = ',work_book_b.sheetnames)

work_book_b.save('users.xlsx')
# Saves a woorkbook on a X path

work_book_b.close()
# Closes a worksheet